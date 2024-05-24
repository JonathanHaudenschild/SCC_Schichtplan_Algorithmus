from datetime import datetime
import random
import math
import numpy as np
import csv
import openpyxl
from openpyxl.styles import PatternFill, Font
from colorhash import ColorHash
import statistics
import time
import concurrent.futures
from functools import partial

FRIENDS_MODE = 2
SHIFT_RANKING_MODE = 0
OFF_DAYS_MODE = 1


# Costs Factor
# Adjust to increase or decrease a cost factor's influence on the total cost.
EXPERIENCE_FACTOR = 1
ONE_SIDED_GENDER_FACTOR = 1
SHIFT_CATEGORY_FACTOR = 1
OFF_DAY_FACTOR = 200
SHIFT_RANKING_FACTOR = 1
CONSECUTIVE_SHIFT_FACTOR = 1
# These are constants representing different levels of preference for or against working with certain partners.
# Negative values are used for preferred partners (friends), with a larger absolute value indicating a stronger preference.
# Positive values are used for non-preferred partners (enemies), with a larger value indicating a stronger preference against.
FRIEND_FACTOR = 100
ENEMY_FACTOR = 250

NUM_OF_SHIFTS_PER_PERSON = 5  # default number of shifts per person

total_obstructed_people = 0


##############
# Excel Crewliste einlesen
##############
def process_excel(file_path):
    people_column_names = [
        "Namen",
        "Spitznamen",
        "Schichtanzahl",
        "Schichtart",
        "Schichtpräferenz",
        "Freunde",
        "Feinde",
        "Freie Schichten",
        "Nicht Verfügbar",
        "Geschlecht",
        "Erfahrung",
        "Minimum",
    ]
    shifts_column_names = ["date", "time", "min", "max", "Schichtart"]
    # Read excel file
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    # Get the first sheet of the workbook
    people_ws = workbook["Personen"]
    # Get the second sheet of the workbook
    shifts_ws = workbook["Schichten"]

    # Initialize an empty dictionary to hold column data
    people_raw_data = {}

    # Iterate through each column
    for column in people_ws.iter_cols(min_row=1, values_only=True):
        # Get the column name
        column_name = column[0]
        # Get the column data as a list
        column_values = [cell for cell in column[1:]]
        # Add the column data to the dictionary
        people_raw_data[column_name] = column_values

    name_data = [
        (
            (int(index), nickname)
            if nickname is not None and nickname != ""
            else (int(index), name)
        )
        for index, name, nickname in zip(
            people_raw_data["index"],
            people_raw_data[people_column_names[0]],
            people_raw_data[people_column_names[1]],
        )
        if name is not None and name != ""
    ]
    capacity_data = [
        (int(index), int(capacity))
        for index, capacity in zip(
            people_raw_data["index"], people_raw_data[people_column_names[2]]
        )
        if capacity is not None
    ]

    # This list defines the preferred shift categories of the people.
    # Each tuple consists of a persons index and a shift category constant.
    # For example, a tuple (4, CHECK_IN) means that the person with index 4 prefers to work check-in shifts.
    # check_in_pref_peopleData, shiftsData = [(i, CHECK_IN) for i, check_in in df[shift_category_col].items() if (check_in == "CHECK_IN") or (check_in == 1)]
    preferred_shift_category_list = [
        (int(index), 1 if shift_category == "CHECK_IN" else shift_category)
        for index, shift_category in zip(
            people_raw_data["index"], people_raw_data[people_column_names[3]]
        )
        if shift_category is not None and str(shift_category) != ""
    ]

    # This list defines the shift rankings of the people.
    # Each tuple consists of a person's index and a shift ranking.
    # A shift ranking is a number that indicates the priority of a shift.
    # A shift ranking of 3 is the highest priority, a shift ranking of 2 is the second highest priority, and so on.
    # A shift ranking of 0 means that the person has no preference for any shift.
    shift_preference_data = [
        (int(index), tuple(int(j) for j in shift_ranking.split(",")))
        for index, shift_ranking in zip(
            people_raw_data["index"], people_raw_data[people_column_names[4]]
        )
        if shift_ranking is not None and str(shift_ranking) != ""
    ]

    # This list defines the preference of people to work together.
    # Each tuple consists of two persons indices and a preference constant.
    friends_data = [
        (int(index), int(float(j)), -1)
        for index, friends in zip(
            people_raw_data["index"], people_raw_data[people_column_names[5]]
        )
        if friends is not None
        for j in str(friends).split(",")
        if j.strip() != ""
    ]
    enemies_data = [
        (int(index), int(float(j)), 1)
        for index, enemies in zip(
            people_raw_data["index"], people_raw_data[people_column_names[6]]
        )
        if enemies is not None
        for j in str(enemies).split(",")
        if j.strip() != ""
    ]
    preference_data = friends_data + enemies_data

    # Process off shifts
    # This list defines the off shifts of the people.
    # Each tuple consists of a person's index and a shift index.
    off_shifts_data = [
        (int(index), tuple(int(val) for val in str(off_day).split(",")))
        for index, off_day in zip(
            people_raw_data["index"], people_raw_data[people_column_names[7]]
        )
        if off_day is not None and str(off_day) != ""
    ]
    # Process unavailable shifts
    # This list defines the unavailable shifts of the people.
    # Each tuple consists of a person's index and a shift index.
    unavailability_data = [
        (int(index), tuple(int(val) for val in str(unavailable).split(",")))
        for index, unavailable in zip(
            people_raw_data["index"], people_raw_data[people_column_names[8]]
        )
        if unavailable is not None and str(unavailable) != ""
    ]
    # This list defines the gender of each person.
    # Each tuple consists of a person index and a gender category.
    gender_data = [
        (int(index), int(gender))
        for index, gender in zip(
            people_raw_data["index"], people_raw_data[people_column_names[9]]
        )
        if gender is not None
    ]
    # This list defines the experience level of each person.
    # Each tuple consists of a person index and a experience level.
    experience_data = [
        (int(index), int(experience))
        for index, experience in zip(
            people_raw_data["index"], people_raw_data[people_column_names[10]]
        )
        if experience is not None
    ]

    # This list defines the minimum number of shifts that are between their shifts
    # Each tuple consists of a person index and a minimum number of shifts
    minimum_data = [
        (int(index), int(minimum))
        for index, minimum in zip(
            people_raw_data["index"], people_raw_data[people_column_names[11]]
        )
        if minimum is not None
    ]

    people_data = {
        "name_data": name_data,
        "capacity_data": capacity_data,
        "preferred_shift_category_data": preferred_shift_category_list,
        "preference_data": preference_data,
        "off_shifts_data": off_shifts_data,
        "unavailability_data": unavailability_data,
        "shift_preference_data": shift_preference_data,
        "gender_data": gender_data,
        "experience_data": experience_data,
        "minimum_data": minimum_data,
    }

    # Initialize an empty dictionary to hold column data
    shifts_raw_data = {}

    # Iterate through each column
    for column in shifts_ws.iter_cols(min_row=1, values_only=True):
        # Get the column name
        column_name = column[0]
        # Get the column data as a list
        column_values = [cell for cell in column[1:]]
        # Add the column data to the dictionary
        shifts_raw_data[column_name] = column_values

    # This list specifies the dates for which the schedule is being generated.
    # Each tuple represents a date's index, the date in string format, and the indices of the shifts on that date.

    shift_date_data = [
        (int(index), shift_name)
        for index, shift_name in zip(
            shifts_raw_data["index"], shifts_raw_data[shifts_column_names[0]]
        )
        if index is not None and shift_name is not None and shift_name != ""
    ]

    # Creating a list of tuples to represent different work shifts.
    # Each tuple contains an index and a corresponding shift time.
    shift_time_data = []
    seen_shift_times = set()
    for index, shift_time in zip(
        shifts_raw_data["index"], shifts_raw_data[shifts_column_names[1]]
    ):
        if (
            shift_time is not None
            and shift_time != ""
            and shift_time not in seen_shift_times
        ):
            shift_time_data.append((int(index), shift_time))
            seen_shift_times.add(shift_time)

    shift_min_data = [
        (int(index), int(shift_min))
        for index, shift_min in zip(
            shifts_raw_data["index"], shifts_raw_data[shifts_column_names[2]]
        )
        if shift_min is not None
    ]

    shift_max_data = [
        (int(index), int(shift_max))
        for index, shift_max in zip(
            shifts_raw_data["index"], shifts_raw_data[shifts_column_names[3]]
        )
        if shift_max is not None
    ]

    shift_capacity_data = [
        (int(index), (min_capacity, max_capacity))
        for (index, min_capacity, max_capacity) in zip(
            shifts_raw_data["index"], shift_min_data, shift_max_data
        )
    ]

    shift_category_data = [
        (int(index), int(shift_category))
        for index, shift_category in zip(
            shifts_raw_data["index"], shifts_raw_data[shifts_column_names[4]]
        )
        if shift_category is not None and shift_category != ""
    ]

    # Creating a list of tuples to represent shift rankings.
    # Each tuple contains a ranking and a corresponding tuple of shift indices.
    # This assigns a priority or preference order to different shifts
    shift_ranking_list = [
        (
            0,
            (0, 1, 2, 3),
        ),  # The shifts at indices 0, 1, 2, and 3 are assigned a ranking of 1
        (
            1,
            (4, 5, 6, 7, 20, 21, 22, 23),
        ),  # The shifts at indices 4, 5, 6, 7, 20, 21, 22, and 23 are assigned a ranking of 3
        (
            5,
            (8, 9, 10, 11),
        ),  # The shifts at indices 8, 9, 10, and 11 are assigned a ranking of 6
        (
            15,
            (12, 13, 14, 15, 16, 17, 18, 19),
        ),  # The shifts at indices 12, 13, 14, 15, 16, 17, 18, and 19 are assigned a ranking of 9
        # Add more shift rankings here
    ]

    # Creating a list of tuples to represent rankings for different shift types.
    # Each tuple contains a ranking and a corresponding tuple of shift type indices.
    # This assigns a priority  or preference order to different shift types,
    shift_type_ranking_list = [
        (1, (0,)),  # The shift type at index 2 is assigned a ranking of 1
        (3, (1, 3)),  # The shift types at indices 1 and 3 are assigned a ranking of 3
        (9, (2,)),  # The shift type at index 0 is assigned a ranking of 9
        # Add more shift type rankings here
    ]

    shifts_data = {
        "shift_date_data": shift_date_data,
        "shift_time_data": shift_time_data,
        "shift_capacity_data": shift_capacity_data,
        "shift_ranking_data": shift_ranking_list,
        "shift_type_ranking_data": shift_type_ranking_list,
        "shift_category_data": shift_category_data,
    }
    return people_data, shifts_data


# Parameters for the simulated annealing algorithm
initial_temperature = 1000  # initial temperature
cooling_rate = 0.99999  # cooling rate
activate_parallelization = True  # activate parallelization
num_of_parallel_threads = 10  # number of parallel threads

##############################################################################################################
# DO NOT CHANGE ANYTHING BELOW THIS LINE
##############################################################################################################


def generate_initial_solution(
    x, y, shift_capacity_matrix, person_capacity_array, max_attempts=10000
):
    solution = [set() for _ in range(x)]
    attempts = 0
    for person in range(y):
        assigned_shifts = set()
        while (len(assigned_shifts) < person_capacity_array[person]) and (
            attempts < max_attempts
        ):
            shift = random.randint(0, x - 1)
            if (
                len(solution[shift]) < shift_capacity_matrix[1][shift][1]
                and shift not in assigned_shifts
            ):
                # Higher probability of adding a person if the shift has less than the minimum people
                if (
                    len(solution[shift]) < shift_capacity_matrix[0][shift][1]
                    or random.random() < 0.30
                ):
                    solution[shift].add(person)
                    assigned_shifts.add(shift)
            attempts += 1

        if attempts >= max_attempts:
            print(
                "Warning: Failed to generate a valid initial solution after "
                + str(max_attempts)
                + " attempts."
            )
            break

    return solution


def get_neighbor(
    solution,
    unavailability_matrix,
    shift_capacity_matrix,
    minimum_array,
    max_attempts=10000,
):
    attempts = 0
    while attempts < max_attempts:
        index1 = random.randint(0, len(solution) - 1)
        index2 = random.randint(0, len(solution) - 1)
        shift1 = solution[index1].copy()
        shift2 = solution[index2].copy()

        if (
            len(shift1) > shift_capacity_matrix[0][index1][1]
            and len(shift2) < shift_capacity_matrix[0][index2][1]
        ):

            person_to_move = get_random_element(shift1)

            if (
                person_to_move not in shift2
                and consecutive_shifts(solution, index2, person_to_move, minimum_array)
                == 0
                and unavailability(index2, unavailability_matrix, person_to_move) == 0
            ):
                shift1.remove(person_to_move)
                shift2.add(person_to_move)
                new_solution = solution.copy()
                new_solution[index1] = shift1
                new_solution[index2] = shift2
                return new_solution  # The neighbor solution satisfies both hard constraints
        else:
            # Swap people between the shifts if possible
            person1 = get_random_element(shift1)
            person2 = get_random_element(shift2)

            if (
                person1 not in shift2
                and person2 not in shift1
                and consecutive_shifts(solution, index2, person1, minimum_array) == 0
                and consecutive_shifts(solution, index1, person2, minimum_array) == 0
                and unavailability(index2, unavailability_matrix, person1) == 0
                and unavailability(index1, unavailability_matrix, person2) == 0
            ):
                shift1.remove(person1)
                shift1.add(person2)
                shift2.remove(person2)
                shift2.add(person1)
                new_solution = solution.copy()
                new_solution[index1] = shift1
                new_solution[index2] = shift2
                return new_solution  # The neighbor solution satisfies both hard constraints

        attempts += 1
    # Return the original solution if no valid neighbor is found after max_attempts
    print("No valid neighbor found after", max_attempts, "attempts")
    return solution


def get_random_element(s):
    return random.choice(list(s))


def adaptive_cooling_rate(initial_temperature, min_temperature, cooling_factor):
    return max(min_temperature, initial_temperature * cooling_factor)


def acceptance_probability(
    old_cost, new_cost, temperature, deviation_old, deviation_new
):
    if new_cost < old_cost or deviation_new < deviation_old:
        return 1
    else:
        return math.exp(-(abs(new_cost - old_cost) / temperature))


def run_parallel_simulated_annealing(
    num_instances, people_data, shifts_data, *args, **kwargs
):

    print("Num_instances", num_instances)
    best_solutions = []

    # Use a ProcessPoolExecutor to run the simulated annealing function concurrently
    with concurrent.futures.ProcessPoolExecutor() as executor:
        # Call the function with different seeds or initial solutions
        seeds = [random.randint(0, 1000000) for _ in range(num_instances)]
        annealing_function = partial(
            simulated_annealing,
            people_data,
            shifts_data,
            initial_temperature,
            cooling_rate,
            1000,
        )
        for result in executor.map(annealing_function, seeds):
            best_solutions.append(result)

    print("Best solutions found:", best_solutions)
    best_solutions.sort(key=lambda x: x[1])
    return best_solutions[0][0], best_solutions[0][1], best_solutions[0][2]


def simulated_annealing(
    people_data,
    shifts_data,
    initial_temperature,
    cooling_rate,
    max_iterations_without_improvement,
    seed=None,
):
    # If a seed is provided, use it to initialize the random module
    if seed is not None:
        random.seed(seed)
    name_list = people_data["name_array"]
    person_capacity_array = people_data["person_capacity_array"]
    unavailability_matrix = people_data["unavailability_matrix"]
    minimum_array = people_data["minimum_array"]
    shift_capacity_matrix = shifts_data["shift_capacity_matrix"]

    num_of_shifts = len(shifts_data["shift_date_array"])  # total number of shifts
    num_people = len(name_list)  # total number of people
    current_solution = generate_initial_solution(
        num_of_shifts, num_people, shift_capacity_matrix, person_capacity_array
    )
    current_cost = cost_function(current_solution, people_data, shifts_data)
    deviation_individual_cost = statistics.stdev(
        individual_cost(current_solution, people_data, shifts_data).values()
    )

    init_cost = current_cost
    temperature = initial_temperature
    iterations_without_improvement = 0

    total_iterations = math.ceil(
        math.log(1 / initial_temperature) / math.log(cooling_rate)
    )
    start_time = time.time()
    current_iteration = 0
    while (
        temperature > 1
        and iterations_without_improvement < max_iterations_without_improvement
    ):

        new_solution = get_neighbor(
            current_solution,
            unavailability_matrix,
            shift_capacity_matrix,
            minimum_array,
        )
        new_cost = cost_function(new_solution, people_data, shifts_data)
        new_deviation_individual_cost = statistics.stdev(
            individual_cost(new_solution, people_data, shifts_data).values()
        )

        if (
            acceptance_probability(
                current_cost,
                new_cost,
                temperature,
                deviation_individual_cost,
                new_deviation_individual_cost,
            )
            > random.random()
        ):
            current_solution = new_solution
            current_cost = new_cost
            deviation_individual_cost = new_deviation_individual_cost
            iterations_without_improvement = (
                0  # Reset the counter when an improvement is found
            )
        else:
            iterations_without_improvement += 1

        temperature *= cooling_rate
        current_iteration += 1
        if current_iteration % 1000 == 0:
            showProgressIndicator(
                current_iteration, total_iterations, start_time, new_cost, init_cost
            )

    return current_solution, current_cost, init_cost


def create_ranking_array(
    shift_ranking_list, shift_type_ranking_list, num_of_shifts, num_of_shift_types
):
    cost_array = [0] * num_of_shifts

    for cost, shifts in shift_ranking_list:
        for shift_index in shifts:
            # Check if this shift_index matches a shift type
            for type_cost, shifts in shift_type_ranking_list:
                for shift in shifts:
                    if (
                        shift_index % num_of_shift_types == shift
                    ):  # Assuming there are 4 shift types
                        cost_array[shift_index] = type_cost * cost

    return cost_array


def create_total_friends_array(preference_list, num_people):
    total_friends_array = [0] * num_people
    for person1, person2, preference in preference_list:
        if preference < 0:
            total_friends_array[person1] += 1
    return total_friends_array


def create_preference_matrix(preference_list, num_people):
    preference_matrix = [[0 for _ in range(num_people)] for _ in range(num_people)]
    for person1, person2, preference in preference_list:
        preference_matrix[person1][person2] = preference
        preference_matrix[person2][person1] = preference

    return preference_matrix


def create_persons_capacity_array(capacity_list, num_people):
    capacity_array = [NUM_OF_SHIFTS_PER_PERSON] * num_people
    for capacity in capacity_list:
        capacity_array[capacity[0]] = capacity[1]
    return capacity_array


def create_preferred_shift_category_array(pref_shift_category_list, num_people):
    pref_shift_category_array = [0] * num_people
    for pref_shift_category in pref_shift_category_list:
        pref_shift_category_array[pref_shift_category[0]] = pref_shift_category[1]
    return pref_shift_category_array


def create_experience_array(experience_list, num_people):
    experience_array = [1] * num_people
    for experience in experience_list:
        experience_array[experience[0]] = experience[1]
    return experience_array


def create_minimum_array(minimum_list, num_people):
    minimum_array = [0] * num_people
    for minimum in minimum_list:
        minimum_array[minimum[0]] = minimum[1]
    return minimum_array


def create_gender_array(gender_list, num_people):
    gender_array = [0] * num_people
    for gender in gender_list:
        gender_array[gender[0]] = gender[1]
    return gender_array


def create_shift_category_array(shift_category_list, num_of_shifts):
    shift_category_array = [0] * num_of_shifts
    for shift_category in shift_category_list:
        shift_category_array[shift_category[0]] = shift_category[1]
    return shift_category_array


def create_preferred_shift_matrix(pref_shift_list, num_of_shift_types, num_people):
    matrix = [[1] * num_of_shift_types for _ in range(num_people)]
    for person, shift_values in pref_shift_list:
        matrix[person] = list(shift_values)
    return matrix


def create_unavailability_matrix(unavailability_list, num_of_shifts, num_people):
    unavailability_matrix = [
        [0 for _ in range(num_of_shifts)] for _ in range(num_people)
    ]
    for person, shifts in unavailability_list:
        for shift in shifts:
            unavailability_matrix[person][shift] = 1
    return unavailability_matrix


def create_shift_capacity_matrix(shift_capacity_list, num_of_shifts):
    shift_capacity_matrix = [[0 for _ in range(num_of_shifts)] for _ in range(2)]
    for shift, capacity in shift_capacity_list:
        shift_capacity_matrix[0][shift] = capacity[0]
        shift_capacity_matrix[1][shift] = capacity[1]
    return shift_capacity_matrix


def cost_function(solution, people_data, shifts_data):
    num_of_shifts = len(shifts_data["shift_date_array"])

    # Calculate individual costs
    individual_costs = individual_cost(solution, people_data, shifts_data)

    # Calculate the mean and standard deviation of individual costs
    mean_individual_cost = statistics.mean(individual_costs.values())
    deviation_individual_cost = (
        statistics.stdev(individual_costs.values()) if len(individual_costs) > 1 else 0
    )

    # Calculate mixed experience and gender costs
    exp_cost = mixedExperience_cost(
        solution, people_data["experience_array"], num_of_shifts
    )
    gender_cost = mixedGender_cost(solution, people_data["gender_array"], num_of_shifts)

    # Introduce a balance factor to penalize high deviation
    balance_factor = 15  # Adjust this factor as needed
    balance_cost = deviation_individual_cost * balance_factor
    # Total cost combines individual costs, experience cost, gender cost, and balance cost
    total_cost = (
        mean_individual_cost
        + deviation_individual_cost
        + balance_cost
        + gender_cost
        + exp_cost
    )
    return total_cost


def individual_cost(solution, people_data, shifts_data):
    num_people = len(people_data["name_array"])
    individual_costs = {person: 0 for person in range(num_people)}

    # Calculate individual preference costs
    pref_costs = preference_cost(
        solution,
        people_data["preference_matrix"],
        people_data["total_friends"],
        people_data["person_capacity_array"],
    )
    for person, cost in enumerate(pref_costs):
        individual_costs[person] += cost

    # Calculate individual off-day costs
    off_day_costs = offDay_cost(solution, people_data["off_shifts_matrix"])
    for person, cost in enumerate(off_day_costs):
        individual_costs[person] += cost

    # Calculate individual shift ranking costs
    rank_costs = shift_ranking_cost(
        solution,
        shifts_data["ranking_array"],
        people_data["preferred_shift_matrix"],
        people_data["off_shifts_matrix"],
        people_data["minimum_array"],
        num_people,
        len(shifts_data["shift_time_array"]),
        shifts_data["shift_type_array"],
    )
    for person, cost in enumerate(rank_costs):
        individual_costs[person] += cost

    # # Calculate individual shift category costs
    # shift_category_costs = shift_category_com_cost(
    #     solution,
    #     shifts_data["shift_category_array"],
    #     people_data["preferred_shift_category_array"],
    # )
    # for person, cost in enumerate(shift_category_costs):
    #     individual_costs[person] += cost

    # print("Total Cost: ",statistics_costs, end='\r' )
    #
    return individual_costs


def preference_cost(
    solution,
    preference_matrix,
    total_friends,
    person_capacity_array,
    friend_factor=FRIEND_FACTOR,
    enemy_factor=ENEMY_FACTOR,
):
    num_people = len(preference_matrix)
    individual_costs = [0] * num_people
    friend_shift_count = [
        0
    ] * num_people  # Track how often each person works with their friends

    # Count the number of times each person works with their friends
    for shift in solution:
        for person1 in shift:
            for person2 in shift:
                if person1 != person2 and preference_matrix[person1][person2] < 0:
                    friend_shift_count[person1] += 1

    # Calculate the preference cost
    for shift in solution:
        for person1 in shift:
            has_friend_in_shift = False
            for person2 in shift:
                if person1 != person2:
                    if preference_matrix[person1][person2] < 0:
                        has_friend_in_shift = True
                        friend_shift_count[person2] -= 1  # Decrement friend shift count

                    if preference_matrix[person1][person2] > 0:
                        individual_costs[person1] += (
                            preference_matrix[person1][person2] * enemy_factor
                        )

            # Penalize if person1 has friends but none in the shift, and the friend still has shift capacity left
            if total_friends[person1] > 0 and not has_friend_in_shift:
                # Check if any friend still has shift capacity left
                for friend, capacity in enumerate(person_capacity_array):
                    if (
                        preference_matrix[person1][friend] < 0
                        and capacity > friend_shift_count[friend]
                    ):
                        individual_costs[person1] += friend_factor
                        break  # Only penalize once per shift if any friend has capacity left

    return individual_costs


def offDay_cost(solution, unavailability_matrix, off_day_factor=OFF_DAY_FACTOR):
    # Initialize individual costs for each person
    individual_costs = [0] * len(unavailability_matrix)

    # Iterate through each shift and person in the shift
    for shift_index, shift in enumerate(solution):
        for person in shift:
            # Check if the person is unavailable for this shift
            if unavailability_matrix[person][shift_index]:
                # Set the cost for the person working during their off day
                individual_costs[person] = off_day_factor

    return individual_costs


def shift_ranking_cost(
    solution,
    ranking_array,
    personal_pref_matrix,
    unavailability_matrix,
    minimum_array,
    num_people,
    num_of_shift_types,
    ranking_type_array,
):
    individual_costs = [0] * num_people
    shift_types = [[1] * num_of_shift_types for _ in range(num_people)]
    last_shift_index = [-1] * num_people
    conc_shifts = [0] * num_people

    for shift_index, shift in enumerate(solution):
        shift_type = shift_index % num_of_shift_types
        shift_cost = ranking_array[shift_index]
        for person in shift:
            persons_cost = 0
            # duplicate_factor = shift_types[person][shift_type]

            # persons_cost +=  shift_types[person][shift_type]
            # shift_types[person][shift_type] += ([t[0] for t in ranking_type_array if shift_type in t[1]][0] * personal_pref_matrix[person][shift_type])
            persons_cost += (
                [t[0] for t in ranking_type_array if shift_type in t[1]][0]
                * personal_pref_matrix[person][shift_type] * shift_cost
            
            )
            if last_shift_index[person] != -1:
                shift_diff = shift_index - last_shift_index[person]

                if conc_shifts[person] > 1:
                    persons_cost += CONSECUTIVE_SHIFT_FACTOR * conc_shifts[person]

                if shift_diff < (minimum_array[person]):
                    conc_shifts[person] += 3
                else:
                    conc_shifts[person] = 0

            last_shift_index[person] = shift_index
            individual_costs[person] += persons_cost

    return individual_costs


def shift_category_com_cost(
    solution,
    shift_category_array,
    pref_shift_category_array,
    shift_category_factor=SHIFT_CATEGORY_FACTOR,
):
    num_people = len(pref_shift_category_array)
    individual_costs = [0] * num_people

    # Count the number of mismatches for each person
    mismatches = [0] * num_people
    for shift_index, shift in enumerate(solution):
        for person in shift:
            if shift_category_array[shift_index] != pref_shift_category_array[person]:
                if shift_category_array[shift_index] > 0:
                    mismatches[person] += 1

    # Calculate the cost based on the number of mismatches
    for person in range(num_people):
        if mismatches[person] > 0:
            individual_costs[person] = shift_category_factor * (
                2 ** (mismatches[person] - 1)
            )  # Exponential increase

    return individual_costs


def mixedExperience_cost(solution, experience_array, num_of_shifts):
    total_cost = 0
    shift_experience = [0] * num_of_shifts

    # Calculate the total experience for each shift
    for shift_index, shift in enumerate(solution):
        total_experience = sum(experience_array[person] for person in shift)
        shift_experience[shift_index] = total_experience

    # Calculate the mean and standard deviation of the shift experiences
    if (
        len(shift_experience) > 1
    ):  # Ensure there are multiple shifts to avoid zero division
        deviation = statistics.stdev(shift_experience)
        mean = statistics.mean(shift_experience)
    else:
        deviation = 0
        mean = shift_experience[0] if shift_experience else 0

    # Calculate the total cost
    # Higher deviation increases the cost, and higher mean decreases the cost
    total_cost = (deviation * EXPERIENCE_FACTOR) / (
        mean if mean != 0 else 1
    )  # Prevent division by zero
    return total_cost


def mixedGender_cost(solution, gender_array, num_of_shifts, one_sided_gender_factor=5):
    shift_gender = []

    # Calculate the gender mix for each shift
    for shift_index, shift in enumerate(solution):
        if len(shift) > 0:
            gender_mix = sum(gender_array[person] for person in shift)
            normalized_gender_mix = gender_mix / len(
                shift
            )  # Normalize by the number of people in the shift
            shift_gender.append(normalized_gender_mix)
        else:
            shift_gender.append(0)  # Handle empty shifts

    # Calculate the standard deviation and mean of the gender mix
    if len(shift_gender) > 1:  # Ensure there are multiple shifts to avoid zero division
        deviation = statistics.stdev(shift_gender)
        mean = statistics.mean(shift_gender)
    else:
        deviation = 0
        mean = shift_gender[0] if shift_gender else 0

    total_average = statistics.mean(gender_array)

    # Calculate the total cost
    total_cost = abs(mean - total_average) * one_sided_gender_factor * deviation

    return total_cost


def unavailability(shift_index, unavailability_matrix, person):
    if unavailability_matrix[person][shift_index]:
        return 1
    return 0


def consecutive_shifts(solution, shift_index, person, minimum_array):
    # Check previous and next shifts within the specified range
    for i in range(1, minimum_array[person] + 1):
        if (shift_index - i >= 0 and person in solution[shift_index - i]) or (
            shift_index + i < len(solution) and person in solution[shift_index + i]
        ):
            return 1

    return 0


def replace_numbers_with_names(solution, index_name_list):
    name_solution = []
    for shift in solution:
        name_shift = set()
        for number in shift:
            name = next(name for index, name in index_name_list if index == number)
            name_shift.add(name)
        name_solution.append(name_shift)
    return name_solution


def showProgressIndicator(
    current_iteration, total_iterations, start_time, new_cost, init_cost
):
    progress = current_iteration / total_iterations
    elapsed_time = time.time() - start_time
    remaining_time = elapsed_time / progress - elapsed_time

    # Convert remaining_time to hours, minutes, and seconds
    hours, remainder = divmod(remaining_time, 3600)
    minutes, seconds = divmod(remainder, 60)

    print(
        f"Progress: {progress * 100:.2f}% | Estimated time remaining: {hours:.0f}h {minutes:.0f}m {seconds:.0f}s | Cost Improvement: {round(((init_cost - new_cost) / init_cost) * 100)}%  ",
        end="\r",
    )


def createFile(solution, shift_name_list, dates_list):
    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write shift names as headers
    for col_index, _ in enumerate(solution):
        shift_name = shift_name_list[col_index % len(shift_name_list)][1]

        cell = worksheet.cell(row=1, column=col_index + 2)
        cell.value = col_index
        for dates in dates_list:
            if dates[0] == col_index:
                day_name = dates[1]
                cell1 = worksheet.cell(row=2, column=col_index + 2)
                cell1.value = day_name
        cell2 = worksheet.cell(row=3, column=col_index + 2)
        cell2.value = shift_name

        # Write the peopleData, shiftsData to the worksheet and apply a unique color to each name
    name_colors = {}
    white_font = Font(color="FFFFFF")  # Set font color to white
    dark_font = Font(color="000000")  # Set font color to black
    for row_index, shift in enumerate(solution, start=4):
        cell = worksheet.cell(row=row_index, column=1)
        cell.value = row_index - 3
        cell.font = dark_font
    for col_index, shift in enumerate(solution, start=1):

        for row_index, name in enumerate(shift, start=4):
            cell = worksheet.cell(row=row_index, column=col_index + 1)
            cell.value = name

            # Generate a unique color for each name if not already generated
            if name not in name_colors:
                color = ColorHash(name).hex
                color = "FF" + color[1:]  # Add alpha channel to the hex color
                name_colors[name] = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

            # Apply the unique color to the cell
            cell.fill = name_colors[name]
            cell.font = white_font
    # Get the current time
    now = datetime.now()

    # Format the current time as a string
    now_str = now.strftime("%H-%M-%S")

    # Save the workbook with the formatted time in the filename
    workbook.save(now_str + "_shifts.xlsx")


def transform_data(people_data, shifts_data):
    num_of_shifts = len(shifts_data["shift_date_data"])
    num_people = len(people_data["name_data"])
    num_of_shift_types = len(shifts_data["shift_time_data"])
    preference_matrix = create_preference_matrix(
        people_data["preference_data"], num_people
    )
    people_transformed_data = {
        "name_array": people_data["name_data"],
        "person_capacity_array": create_persons_capacity_array(
            people_data["capacity_data"], num_people
        ),
        "preference_matrix": preference_matrix,
        "unavailability_matrix": create_unavailability_matrix(
            people_data["unavailability_data"], num_of_shifts, num_people
        ),
        "off_shifts_matrix": create_unavailability_matrix(
            people_data["off_shifts_data"], num_of_shifts, num_people
        ),
        "preferred_shift_matrix": create_preferred_shift_matrix(
            people_data["shift_preference_data"], num_of_shift_types, num_people
        ),
        "preferred_shift_category_array": create_preferred_shift_category_array(
            people_data["preferred_shift_category_data"], num_people
        ),
        "gender_array": create_gender_array(people_data["gender_data"], num_people),
        "experience_array": create_experience_array(
            people_data["experience_data"], num_people
        ),
        "minimum_array": create_minimum_array(people_data["minimum_data"], num_people),
        "total_friends": create_total_friends_array(
            people_data["preference_data"], num_people
        ),
    }

    shifts_transformed_data = {
        "shift_date_array": shifts_data["shift_date_data"],
        "shift_time_array": shifts_data["shift_time_data"],
        "shift_capacity_matrix": create_shift_capacity_matrix(
            shifts_data["shift_capacity_data"], num_of_shifts
        ),
        "ranking_array": create_ranking_array(
            shifts_data["shift_ranking_data"],
            shifts_data["shift_type_ranking_data"],
            num_of_shifts,
            num_of_shift_types,
        ),
        "shift_category_array": create_shift_category_array(
            shifts_data["shift_category_data"], num_of_shifts
        ),
        "shift_type_array": shifts_data["shift_type_ranking_data"],
    }
    return people_transformed_data, shifts_transformed_data


def check_person_costs(solution, people_data, shifts_data):
    individual_costs = individual_cost(solution, people_data, shifts_data)

    off_day_costs = offDay_cost(solution, people_data["unavailability_matrix"])
    pref_costs = preference_cost(
        solution,
        people_data["preference_matrix"],
        people_data["total_friends"],
        people_data["person_capacity_array"],
    )
    rank_costs = shift_ranking_cost(
        solution,
        shifts_data["ranking_array"],
        people_data["preferred_shift_matrix"],
        people_data["off_shifts_matrix"],
        people_data["minimum_array"],
        len(people_data["name_array"]),
        len(shifts_data["shift_time_array"]),
        shifts_data["shift_type_array"],
    )
    # shift_category_costs = shift_category_com_cost(
    #     solution,
    #     shifts_data["shift_category_array"],
    #     people_data["preferred_shift_category_array"],
    # )

    for person in range(len(people_data["name_array"])):
        print(f"Person {person}: Total Cost = {individual_costs[person]}")
        print(f"    Preference Cost: {pref_costs[person]}")
        print(f"    Off-Day Cost: {off_day_costs[person]}")
        print(f"    Shift Ranking Cost: {rank_costs[person]}")
        print(f"    Total Cost Breakdown: {individual_costs[person]}")

    return individual_costs


# Main Funktion
if __name__ == "__main__":

    people_data, shifts_data = process_excel("SCC_SCHICHTPLAN_FINAL.xlsx")
    people_transformed_data, shifts_transformed_data = transform_data(
        people_data, shifts_data
    )
    shift_time_list = shifts_transformed_data["shift_time_array"]
    name_list = people_transformed_data["name_array"]
    dates_list = shifts_transformed_data["shift_date_array"]
    if activate_parallelization:
        best_solution, best_cost, init_cost = run_parallel_simulated_annealing(
            num_of_parallel_threads, people_transformed_data, shifts_transformed_data
        )
    else:
        best_solution, best_cost, init_cost = simulated_annealing(
            people_transformed_data,
            shifts_transformed_data,
            initial_temperature,
            cooling_rate,
            max_iterations_without_improvement=1000,
        )

    # Check the cost of each person
    individual_costs = check_person_costs(
        best_solution, people_transformed_data, shifts_transformed_data
    )
    best_solution_with_names = replace_numbers_with_names(best_solution, name_list)
    print(f"Best solution with names: {best_solution_with_names}")
    print(f"Initial cost: {init_cost}")
    print(f"Best cost: {best_cost}")

    createFile(
        [sorted(sublist) for sublist in best_solution_with_names],
        shift_time_list,
        dates_list,
    )
