import openpyxl
from datetime import datetime, time
from openpyxl.styles import PatternFill, Font
from colorhash import ColorHash
from datetime import datetime, timezone


def convert_time(time_str):
    if isinstance(time_str, time):
        return time_str
    elif isinstance(time_str, float):
        time_str = "{:02.0f}:00:00".format(time_str)
    return datetime.strptime(time_str, "%H:%M:%S").time()


def extract_data(data, column_name, data_type, default=None):
    return [
        (
            id,
            (
                data_type(value)
                if value is not None
                else data_type(default) if default is not None else None
            ),
        )
        for id, value in zip(data["id"], data[column_name])
        if id is not None
    ]


def extract_min_max(data, column_min_name, column_max_name, data_type, default=None):
    return [
        (
            id,
            (
                (
                    data_type(min_val)
                    if min_val is not None
                    else data_type(default) if default is not None else None
                ),
                (
                    data_type(max_val)
                    if max_val is not None
                    else data_type(default) if default is not None else None
                ),
            ),
        )
        for id, min_val, max_val in zip(
            data["id"], data[column_min_name], data[column_max_name]
        )
        if id is not None
    ]


def extract_time_frame(data, column_name, data_type):
    return [
        (id, data_type(value))
        for id, value in zip(data["id"], data[column_name])
        if value is not None
    ]


def extract_shift_types(data, column_name):
    def parse_shift(shift_str):
        shift_dict = {}
        for shift in shift_str.strip("()").split("), ("):
            key, values = shift.split(": ")
            key = int(key.strip())
            values = tuple(map(int, values.strip("()").split(",")))
            shift_dict[key] = values
        return shift_dict

    return [
        (id, parse_shift(value))
        for id, value in zip(data["id"], data[column_name])
        if value is not None
    ]


def extract_times(data, column_name):
    def parse_off_time(value):
        # Remove surrounding parentheses and split by "), ("
        periods = value.strip("()").split("), (")
        result = []
        for period in periods:
            # Split by ", " to separate the two date-time strings
            start, end = period.split(", ")
            # Convert the date-time strings to datetime objects
            start_dt = datetime.strptime(start, "%d/%m/%Y %H:%M:%S")
            end_dt = datetime.strptime(end, "%d/%m/%Y %H:%M:%S")
            # Add the tuple of datetime objects to the result list
            result.append((start_dt, end_dt))
        return result

    return [
        (id, parse_off_time(value))
        for id, value in zip(data["id"], data[column_name])
        if value
    ]


def extract_shift_preferences(data, column_name):
    def parse_shift_pref(value):
        # Remove surrounding parentheses and split by "), ("
        shift_prefs = value.strip("()").split("), (")
        result = []
        for shift_pref in shift_prefs:
            # Split by ", " to separate the date-time strings and the value
            period, value = shift_pref.rsplit(", ", 1)
            start, end = period.strip("()").split(", ")

            # Convert the date-time strings to datetime objects
            start_dt = convert_time(start)
            end_dt = convert_time(end)

            # Add the tuple of datetime objects and the value to the result list
            result.append(((start_dt, end_dt), int(value)))
        return result

    return [
        (id, parse_shift_pref(value))
        for id, value in zip(data["id"], data[column_name])
        if value
    ]


def extract_preferences(data, friends_column_name, enemies_column_name):
    def parse_values(values, flag):
        if isinstance(values, str):
            values = values.split(",")
        elif isinstance(values, (int, float)):
            values = [values]
        return [
            (int(float(j)), flag)
            for j in values
            if isinstance(j, (int, float))
            or (isinstance(j, str) and j.strip().replace(".", "", 1).isdigit())
        ]

    result = []
    for id, friends, enemies in zip(data["id"], data[friends_column_name], data[enemies_column_name]):
        preferences = []
        if friends:
            preferences.extend(parse_values(friends, -1))
        if enemies:
            preferences.extend(parse_values(enemies, 1))
        result.append((id, preferences))

    return result


def extract_availability_data(data, column_name):
    return [
        (id, tuple(int(val) for val in value.split(",")))
        for id, value in zip(data["id"], data[column_name])
        if value
    ]


def process_shifts_data(file_path):

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    shifts_ws = workbook["Schichten"]

    shifts_raw_data = {
        column[0]: [cell for cell in column[1:]]
        for column in shifts_ws.iter_cols(min_row=1, values_only=True)
    }

    shift_time_data = [
        (
            id,
            (
                datetime.strptime(f"{start}", "%Y-%m-%d %H:%M:%S"),
                datetime.strptime(f"{end}", "%Y-%m-%d %H:%M:%S"),
            ),
        )
        for id, start, end in zip(
            shifts_raw_data["id"],
            shifts_raw_data["start"],
            shifts_raw_data["end"],
        )
        if start and end
    ]

    shift_capacity_data = [
        (id, (int(min_val), int(max_val)))
        for id, min_val, max_val in zip(
            shifts_raw_data["id"], shifts_raw_data["min"], shifts_raw_data["max"]
        )
        if min_val is not None and max_val is not None
    ]

    shift_type_data = extract_data(shifts_raw_data, "shift-type", int)
    restrict_shift_type = extract_data(shifts_raw_data, "restrict-shift-type", bool)
    shift_cost_data = extract_data(shifts_raw_data, "cost", int)
    shift_priority_data = extract_data(shifts_raw_data, "priority", int)

    return {
        "shift_time_data": shift_time_data,
        "shift_capacity_data": shift_capacity_data,
        "shift_type_data": shift_type_data,
        "restrict_shift_type_data": restrict_shift_type,
        "shift_cost_data": shift_cost_data,
        "shift_priority_data": shift_priority_data,
    }


def process_people_data(file_path):

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    people_ws = workbook["Personen"]

    people_raw_data = {
        column[0]: [cell for cell in column[1:]]
        for column in people_ws.iter_cols(min_row=1, values_only=True)
    }

    name_data = [
        (id, nickname if nickname else name)
        for id, name, nickname in zip(
            people_raw_data["id"], people_raw_data["firstname"], people_raw_data["nickname"]
        )
        if name
    ]

    capacity_data = extract_min_max(people_raw_data, "min-shifts", "max-shifts", int, 0)
    gender_data = extract_data(people_raw_data, "gender", int)
    experience_data = extract_data(people_raw_data, "experience", int)
    shift_types_data = extract_shift_types(people_raw_data, "shift-types")
    minimum_break_data = extract_data(
        people_raw_data, "minimum_break", convert_time, "12:00:00"
    )

    day_off_data = extract_times(people_raw_data, "day_off")
    unavailability_data = extract_times(people_raw_data, "unavailability_times")
    mandatory_data = extract_times(people_raw_data, "mandatory_times")

    preference_data = extract_preferences(people_raw_data, "friends", "enemies")
    shift_preference_data = extract_shift_preferences(
        people_raw_data, "shift-preference"
    )

    return {
        "name_data": name_data,
        "capacity_data": capacity_data,
        "shift_types_data": shift_types_data,
        "day_off_data": day_off_data,
        "unavailability_data": unavailability_data,
        "minimum_break_data": minimum_break_data,
        "preference_data": preference_data,
        "shift_preference_data": shift_preference_data,
        "gender_data": gender_data,
        "experience_data": experience_data,
        "mandatory_data": mandatory_data,
    }



  

def timestamp_to_datetime(timestamp):
    """
    Convert a singular integer representing the total seconds since the epoch to a datetime object.
    """
    return datetime.fromtimestamp(timestamp, tz=timezone.utc)


def create_file(
    best_schedule,
    total_cost_breakdown,
    people_data,
    shifts_data,
    cost_details,
):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Shifts"
    name_list = people_data["name_dict"]

    col_index = 1
    # Write shift names as headers
    for shift in best_schedule:
        shift_name = shifts_data["shift_time_dict"].get(shift)
        cell = worksheet.cell(row=1, column=col_index * 2 + 2)
        cell.value = col_index

        cell2 = worksheet.cell(row=3, column=col_index * 2 + 2)
        cell2.value = (
            timestamp_to_datetime(shift_name[0]).strftime("%d/%m/%Y")
            + " - "
            + timestamp_to_datetime(shift_name[1]).strftime("%d/%m/%Y")
        )

        col_index += 1

    name_colors = {}
    white_font = Font(color="FFFFFF")
    dark_font = Font(color="000000")

    # Track the maximum row needed to accommodate the longest shift
    max_row = 4

    col_index = 4

    for shift in best_schedule:

        shift_with_names = [
            name_list.get(person_id, "n/a") for person_id in best_schedule[shift]
        ]

        row_index = 4

        # Sort and write shift names for the shift type
        names = sorted(shift_with_names)
        for name in names:
            row_index += 1
            name_cell = worksheet.cell(row=row_index, column=col_index)
            name_cell.value = name
            if name not in name_colors:
                color = ColorHash(name).hex
                color = "FF" + color[1:]
                name_colors[name] = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
            name_cell.fill = name_colors[name]
            name_cell.font = white_font

            cost_cell = worksheet.cell(row=row_index, column=col_index + 1)
            cost_cell.fill = name_colors[name]
            cost_cell.value = total_cost_breakdown.get(name, 0)
            cost_cell.font = white_font

        # Update the max row if necessary
        if row_index > max_row:
            max_row = row_index

        col_index += 2

    # # Add row numbers for the shifts
    for row in range(5, max_row):
        cell = worksheet.cell(row=row, column=1)
        cell.value = row - 4
        cell.font = dark_font

    # Create a new worksheet for the cost details
    cost_details_sheet = workbook.create_sheet(title="Cost Details")

    # Write the cost details to the new worksheet, split at ":" and "="
    cost_details = cost_details.split("\n")
    for row_index, line in enumerate(cost_details, start=1):
        parts = [part.strip() for part in line.replace("=", ":").split(":")]
        for col_index, part in enumerate(parts, start=1):
            cell = cost_details_sheet.cell(row=row_index, column=col_index)
            # Try to convert to integer or float
            try:
                if "." in part:
                    cell.value = float(part)
                else:
                    cell.value = int(part)
            except ValueError:
                cell.value = part

    # Get the current time
    now = datetime.now()
    now_str = now.strftime("%H-%M-%S")

    # Save the workbook with the formatted time in the filename
    workbook.save(now_str + "_shifts.xlsx")


def convert_names_to_indices(shift_data, name_list):
    converted_shift_data = {}
    for shift_index in shift_data:
        converted_shift_data[shift_index] = {}
        for shift_type in shift_data[shift_index]:
            converted_shift_data[shift_index][shift_type] = [
                next(index for index, name in name_list if name == person_name)
                for person_name in shift_data[shift_index][shift_type]
            ]
    return converted_shift_data


def reconstruct_solution_matrix(converted_shift_data, num_shifts, num_shift_types):
    solution = [
        {shift_type: set() for shift_type in range(num_shift_types)}
        for _ in range(num_shifts)
    ]
    for shift_index in converted_shift_data:
        for shift_type in converted_shift_data[shift_index]:
            solution[shift_index][shift_type] = set(
                converted_shift_data[shift_index][shift_type]
            )
    return solution


def load_excel_and_create_solution(file_path, dates_list, name_list, shift_types):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook["Shifts_RAW"]

    num_shifts = len(dates_list)
    # Read the shift data from the worksheet
    solution_matrix = [
        {shift_type: set() for shift_type in shift_types} for _ in range(num_shifts)
    ]

    current_shift_type = None
    shift_index = -1

    for col in range(1, worksheet.max_column + 1, 2):
        shift_index += 1
        current_shift_type = None
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=col).value

            if cell_value is None:
                continue

            if cell_value in shift_types:
                current_shift_type = cell_value
            elif current_shift_type is not None:
                name = cell_value
                index = next(
                    (
                        i
                        for i, name_tuple in enumerate(name_list)
                        if name_tuple[1] == name
                    ),
                    None,
                )
                if index is not None:
                    solution_matrix[shift_index][current_shift_type].add(index)

    return solution_matrix
