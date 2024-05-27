import random
import time
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from colorhash import ColorHash
import math


def replace_numbers_with_names(solution, index_name_list):
    name_solution = []
    for shift in solution:
        name_shift = set()
        for number in shift:
            name = next(name for index, name in index_name_list if index == number)
            name_shift.add(name)
        name_solution.append(name_shift)
    return name_solution

def showProgressIndicator(
    current_iteration, total_iterations, start_time, new_cost, init_cost
):
    progress = current_iteration / total_iterations
    elapsed_time = time.time() - start_time
    remaining_time = elapsed_time / progress - elapsed_time

    hours, remainder = divmod(remaining_time, 3600)
    minutes, seconds = divmod(remainder, 60)

    print(
        f"Progress: {progress * 100:.2f}% | Estimated time remaining: {hours:.0f}h {minutes:.0f}m {seconds:.0f}s | Cost Improvement: {round(((init_cost - new_cost) / init_cost) * 100)}%  | Current Cost: {new_cost:.1f}",
        end="\r",
    )



def createFile(best_solution, name_list, individual_costs, shift_name_list, dates_list):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write shift names as headers
    for col_index, _ in enumerate(best_solution):
        shift_name = shift_name_list[col_index % len(shift_name_list)][1]
        cell = worksheet.cell(row=1, column=col_index + 2)
        cell.value = col_index
        for dates in dates_list:
            if dates[0] == col_index:
                day_name = dates[1]
                cell1 = worksheet.cell(row=2, column=col_index + 2)
                cell1.value = day_name
        cell2 = worksheet.cell(row=3, column=col_index + 2)
        cell2.value = shift_name

    name_colors = {}
    white_font = Font(color="FFFFFF")
    dark_font = Font(color="000000")

    # Track the maximum row needed to accommodate the longest shift (normal + sv)
    max_row = 4

    for col_index, shift in enumerate(best_solution, start=1):
        row_index = 4

        # Write "normal" header
        normal_header_cell = worksheet.cell(row=row_index, column=col_index + 1)
        normal_header_cell.value = "Normal"
        normal_header_cell.font = dark_font
        row_index += 1

        # Sort and write normal shift names
        normal_shift_names = sorted(
            [
                next(name for index, name in name_list if index == person)
                for person in shift
            ]
        )
        for name in normal_shift_names:
            index = next(
                index for index, name_list_name in name_list if name_list_name == name
            )
            cell = worksheet.cell(row=row_index, column=col_index + 1)
            cell.value = f"{name} ({individual_costs[index]})"
            if name not in name_colors:
                color = ColorHash(name).hex
                color = "FF" + color[1:]
                name_colors[name] = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
            cell.fill = name_colors[name]
            cell.font = white_font
            row_index += 1

        # Update the max row if necessary
        if row_index > max_row:
            max_row = row_index

    # Add row numbers for the shifts
    for row in range(5, max_row):
        cell = worksheet.cell(row=row, column=1)
        cell.value = row - 4
        cell.font = dark_font

    # Get the current time
    now = datetime.now()
    now_str = now.strftime("%H-%M-%S")

    # Save the workbook with the formatted time in the filename
    workbook.save(now_str + "_shifts.xlsx")
